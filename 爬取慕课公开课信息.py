import time
import requests
import re
import asyncio
import aiohttp
from aiocsv import AsyncWriter
import aiofiles
import os
from openpyxl import Workbook, load_workbook
import datetime
import json

fake_ug = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36'
}


def get_school_inf_from_excel(path='schools_inf.xlsx'):
    """
    从固定格式的excel文件中读取学校的信息存入列表中
    :param path:
    :return:返回的是列表，列表每一项是一个字典，每个字典存放学校的school_name和href
    """
    work_book = load_workbook(path)
    wook_sheet = work_book['Sheet']
    school_list = []
    for index_row in range(1, wook_sheet.max_row + 1):
        school_inf = {'school_name': wook_sheet.cell(row=index_row, column=1).value,
                      'school_id': wook_sheet.cell(row=index_row, column=4).value,
                      'school_href': wook_sheet.cell(row=index_row, column=2).value}
        school_list.append(school_inf)
    return school_list


async def aio_get_school_id_from_school_domain(school, session):
    """
    请求学校首页源码，从中匹配出school_id
    :param session: 维护所有请求的会话
    :param school:
    :return:
    """
    school_domain = school['href']
    # 首先拼接url
    url = f'https://www.icourse163.org{school_domain}'
    response = await session.get(url)
    assert response.status == 200  # 这是新学的一招，感觉很不错
    # response.encoding = 'utf-8'  # 经测试编码设置应该可以在session进行设置
    content = await response.text()
    school_id = re.search(
        r'window.schoolId = "(?P<school_id>.*?)";', content).group('school_id')
    school['school_id'] = school_id
    # 这里设置一个异步协程检查点
    # print(f'现在返回了{school["school_name"]}的信息:{school_id}')
    # 检查完毕，确实是异步的请求，速度感受也很快
    # FIXME: 这里的返回值似乎无效
    # return school_id


def get_school_id_from_school_domain_manager(schools_inf):
    """
    这个是获取school_id的管理器函数
    :param schools_inf:
    :return:
    """
    # 获取事件循环
    loop = asyncio.get_event_loop()
    # 创建tasks
    tasks = []

    # 创建请求会话
    async def request():
        async with aiohttp.ClientSession(headers=fake_ug) as session:
            session.encoding = 'utf-8'
            for school in schools_inf:
                # school_domain = school['href']  # 多一步中间变量传递没有坏处
                # 直接传递school,方便插入信息
                task = loop.create_task(
                    aio_get_school_id_from_school_domain(school, session=session))
                tasks.append(task)
            await asyncio.wait(tasks)

    loop.run_until_complete(request())
    # 现在获取了学校的信息，将信息返回，调用其他函数保存
    print(schools_inf)
    return schools_inf


def save_school_inf_to_excel(schools_inf, path='schools_inf.xlsx'):
    """
    将新解析的数据从迭代器中读出来，存入本地excel表格进行持久化保持
    :param schools_inf:
    :param path:excel的保存路径
    :return:
    """
    work_book = Workbook()
    work_sheet = work_book.active
    for school in schools_inf:
        school_inf = [school['school_name'],
                      school['href'],
                      school['img'],
                      school['school_id']]
        work_sheet.append(school_inf)
    work_book.save(path)


def parse_schools_domain():
    """
    请求所有学校主页的url,然后解析，并返回包含所有学校的信息
    :return:返回包含所有学校信息的列表
    """
    domain_url = 'https://www.icourse163.org/university/view/all.htm#/'
    response = requests.get(domain_url, headers=fake_ug)
    response.encoding = 'utf-8'
    response.close()
    # 获取到主页的源码后，将数据通过正则表达式记录下来
    # 首先匹配到所有的学习的div
    schools_inf = re.search(
        r'<div class="u-usitys f-cb">.*?</div>', response.text, re.S)
    if schools_inf:
        # 此时成功获取到数据
        # print(schools_inf.group())
        reg_school = re.compile(r'<a class="u-usity f-fl" href="(?P<href>.*?)" target="_blank">.*?'
                                r'src="(?P<img>.*?)".*?'
                                r'alt="(?P<school_name>.*?)"', re.S)
        school_iter = re.finditer(reg_school, schools_inf.group())
        # 数据匹配完毕
        # 将数据迭代器转换成列表更加安全
        schools_inf = []
        for school in school_iter:
            school_inf = {'school_name': school.group('school_name'), 'href': school.group('href'),
                          'img': school.group('img')}
            schools_inf.append(school_inf)
        return schools_inf
    else:
        raise Exception(f'学校的信息没有请求到,response={response}')


def save_school_inf_to_excel_manager(path='schools_inf.xlsx'):
    """
    负责请求与保存数据的管理器函数
    当存储各个学校的基本信息表格缺失时候调用该函数
    :param path:存放表格的路径
    :return:
    """
    schools_inf = parse_schools_domain()
    # 将学校信息写入数据表格中存储
    schools_inf = get_school_id_from_school_domain_manager(schools_inf)
    save_school_inf_to_excel(schools_inf, path=path)


async def aio_save_json_file(json_list, filename, lock):
    async with lock:
        async with aiofiles.open(filename, mode='a', newline='') as afp:
            writer = AsyncWriter(afp)
            data_list = json_list['result']['list']
            # data_list是一个列表，里面每一项是一个课程字典
            if data_list:  # 防止无数据的情况
                for course in data_list:
                    # 下面进行费力气的类型检查
                    course_dict = course.items()
                    for key, value in course_dict:
                        if isinstance(value, type(None)):
                            course[key] = 'null'
                        if isinstance(value, list):
                            course[key] = json.dumps(value[0])
                        if not isinstance(value, (str, int, float, type(None), list)):
                            course[key] = f'typeError-{key}'
                    # 最终存的是课程的值
                    class_inf = list(course.values())
                    await writer.writerow(class_inf)


async def aio_get_school_inf(session, csrfKey, school, lock, filename, page, index=None):
    school_id = school['school_id']
    response = await session.post(
        url=f'https://www.icourse163.org/web/j/courseBean.getCourseListBySchoolId.rpc?csrfKey={csrfKey}',
        data={
            'schoolId': school_id,
            'p': page,
            'psize': 20,
            'type': 1,
            'courseStatus': 30  # 这个参数的含义不清楚
        }
    )
    assert response.status == 200
    answer = await response.json()
    # print(answer)
    await aio_save_json_file(answer, filename, lock)
    # 打印一下任务完成状态
    print(f'位于任务队列第{index}位的{school["school_name"]}完成第{page}页请求')
    # 从数据包中拿出来页码
    pages = answer['result']['query']['totlePageCount']
    # 将页码保存到school的列表参数中
    school['pages'] = pages
    return answer


def prepare_output_file(filename):
    # 创建一个新的表格
    if os.path.exists(filename):
        os.remove(filename)
    with open(filename, mode='w+') as f:
        f.writelines(
            'id,name,bigPhoto,enrollCount,teacherId,teacherName,schoolId,schoolName,schoolSN,currentTermId,startTime,endTime,mode,channel,firstPublishTime,currentTermPubblishStatus,webVisible,closeVisableStatus,isEnroll,evaluateAvgScore,tags\n')


async def aio_get_schools_first_page_pages(schools, filename):
    # 首先创建一个会话，用于为400多个学校进行请求
    # timeout = aiohttp.ClientTimeout(total=600)  # 增加超时时间
    # connector = aiohttp.TCPConnector(limit=40)  # 将并发数量降低
    # 创建协程锁
    lock = asyncio.Lock()
    async with aiohttp.ClientSession(headers=fake_ug) as session:
        # 然后去请求html后自动获取两个cookies
        response = await session.get('https://www.icourse163.org' + schools[0]['school_href'])
        # 然后检查一下返回状态
        assert response.status == 200
        csrfKey = response.cookies['NTESSTUDYSI'].value
        # 准备好会话请求参数了，下面对所有的学校进行请求
        # TODO 确认async for 的工作原理
        tasks = []
        for index, school in enumerate(schools):
            tasks.append(
                asyncio.create_task(
                    aio_get_school_inf(
                        session, csrfKey, school, lock, filename, page=1, index=index)
                )
            )
        await asyncio.wait(tasks)


async def aio_get_schools_other_pages(schools, filename):
    # 首先创建一个会话，用于为400多个学校进行请求
    # timeout = aiohttp.ClientTimeout(total=600)  # 增加超时时间
    # connector = aiohttp.TCPConnector(limit=40)  # 将并发数量降低
    # 创建协程锁
    lock = asyncio.Lock()
    async with aiohttp.ClientSession(headers=fake_ug) as session:
        # 然后去请求html后自动获取两个cookies
        response = await session.get('https://www.icourse163.org' + schools[0]['school_href'])
        # 然后检查一下返回状态
        assert response.status == 200
        csrfKey = response.cookies['NTESSTUDYSI'].value
        # 准备好会话请求参数了，下面对所有的学校进行请求
        tasks = []
        index = 1  # 设置位置标记
        # FIXME 判断index是否需要加锁
        for school in schools:
            for page in range(2, school['pages']):
                tasks.append(
                    asyncio.create_task(
                        aio_get_school_inf(
                            session, csrfKey, school, lock, filename, page=page, index=index)
                    )
                )
                index = index + 1
        await asyncio.wait(tasks)


def main():
    """主函数，完整程序的起点"""
    # 定义常量
    XLSX_IN_FILE_NAME = 'schools_inf.xlsx'
    CSV_OUT_FILE_NAME = f'course_{datetime.date.today().__format__("%Y_%m_%d")}.csv'
    if not os.path.exists(XLSX_IN_FILE_NAME):
        save_school_inf_to_excel_manager(XLSX_IN_FILE_NAME)
    # 此时配置文件已经存在
    # 每一项是一个字典，每个字典包含school_name和school_id
    school_list = get_school_inf_from_excel(XLSX_IN_FILE_NAME)
    # 此时学校的信息已经获取，接下来进入每一个学校进行分析网页
    prepare_output_file(CSV_OUT_FILE_NAME)
    # 创建事件循环
    loop = asyncio.get_event_loop()
    loop.run_until_complete(aio_get_schools_first_page_pages(
        school_list, CSV_OUT_FILE_NAME))
    print(school_list)
    # 接下来对剩余的课程信息进行获取
    loop.run_until_complete(aio_get_schools_other_pages(
        school_list, CSV_OUT_FILE_NAME))
    print('所有任务正常完成')


def test():
    """
    用来测试的函数
    :return:
    """
    school_id = 13001
    school_url = 'https://www.icourse163.org/university/PKU'
    filename = 'course.csv'
    # 创建一个新的表格
    if os.path.exists(filename):
        os.remove(filename)


if __name__ == '__main__':
    start = time.perf_counter()
    main()
    end = time.perf_counter()
    time_used = end - start
    print(f'耗时:{time_used}')
    # test()
